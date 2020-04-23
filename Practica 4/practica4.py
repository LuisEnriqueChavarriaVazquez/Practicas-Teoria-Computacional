#Practica 4 Ricardo Alberto Machorro Vences
#Esta es la practica4 use la libreria openpyxl para usar excel por que es mas facil asi, se tiene que descargar nadamas
from openpyxl import load_workbook

#Un auotomata convexo es el grupo de estados de un automata que ya se verifico que se puede llegar a ellos desde le estado inicial

#La tabla en excel tiene la siguiente estructura
#
#   el simbolo " → " es para el estado incial
#   el simbolo " * " es para estados finales
#
#
#               |     δ     |     entrada 1   |    entrada 2    |    entrada 3    | ....... |     entrada n
#               --------------------------------------------------------------------------------------------
#               | →estado 1 | estado salida 1 | estado salida 2 | estado salida 3 |.........| estado salida n
#               --------------------------------------------------------------------------------------------
#               | *estado 2 | estado salida 1 | estado salida 2 | estado salida 3 |.........| estado salida n
#               ---------------------------------------------------------------------------------------------
#               |  estado 3 | estado salida 1 | estado salida 2 | estado salida 3 |.........| estado salida n
#               ----------------------------------------------------------------------------------------------
#               ..............................................................................................
#               |  estado m | estado salida 1 | estado salida 2 | estado salida 3 |.........| estado salida n
#

primerEstadoInicial=""#Aqui se da el primer estado inicial antes de pasarlo al automata convexo
listaEstadosFinalesIniciales=[]#Aqui se da los estados finales antes de pasarlo al automato convexo
listaEstadosIniciales=[]#Aqui se da los estados antes de pasarlos al automato convexo
listaEstadosDesechadosConvexo=set()#Aqui se da los estados que no llegaron al automato convexo
listaEstadosAutomataConvexo=set()#Aqui esta los estados del automata convexo
listaEstadosFinalesConvexo=set()#Aqui esta los estados finales del auotomata convexo

def verprimerEstadoInicial(posibleprimerEstadoInicial):#Checa si al revisar la tabla un estado es inicial
    if(posibleprimerEstadoInicial[0]=="→"):
        return True
    else:
        return False

def verEstadosFinal(posibleEstadoFinal):#Checa si al revisar la tabla un estado es final
    if(posibleEstadoFinal[0]=="*"):
        return True
    else:
        return False

def estadosConvexo(estadoInicial,listaEstadosFinales,listaEstados,sheetExcel):#Esta funcion esta de mas
    final=False
    listaEstVerificados=set()
    listaEstVerificados.add(estadoInicial)
    listaEstadosEsperaAnt=set()
    listaEstadosEsperaAnt.add(estadoInicial)
    listaEstadosEsperaNuevos=set()
    while(final==False):
        for row in sheetExcel.iter_rows(min_row=2):
            if(row[0].value!=None):
                valorCelda=""
                if(verprimerEstadoInicial(row[0].value) or verEstadosFinal(row[0].value)):
                    valorCelda=row[0].value[1:len(row[0].value)]
                else:
                    valorCelda=row[0].value
                if(valorCelda in listaEstadosEsperaAnt):
                    for i in range(1,sheetExcel.max_column):
                        listaEstVerificados.add(row[i].value)
                        listaEstadosEsperaNuevos.add(row[i].value)
            else:
                break
                break      
        listaEstadosEsperaNuevos=listaEstadosEsperaNuevos-listaEstadosEsperaAnt
        if (len(listaEstadosEsperaNuevos)==0):
            final=True
        else:
            listaEstadosEsperaAnt=listaEstadosEsperaNuevos
            listaEstadosEsperaNuevos.clear()
    print(listaEstVerificados)

#Esta funcion regresa los estados del automataConvexo, es decir aquellos que ya se sabe que se pude llegar desde el estado inicial
#el sheetExcel es el objeto que permite leer datos de excel
def estadosConvexosNuevo(estadoInicial,listaEstadosFinales,listaEstados,sheetExcel):
    final=False
    listaEstadosVerificados=set()
    listaEstadosBusTerminada=set()
    listaEstadosVerificados.add(estadoInicial)
    listaEstadosBusPendienteAhora=set()
    listaEstadosBusPendienteAhora.add(estadoInicial)
    listaEstadosBusPendienteNuevos=set()
    while(final==False):
        for row in sheetExcel.iter_rows(min_row=2):
            if(row[0].value!=None):
                valorCelda=""
            if(verprimerEstadoInicial(row[0].value) or verEstadosFinal(row[0].value)):
                valorCelda=row[0].value[1:len(row[0].value)]
            else:
                valorCelda=row[0].value
            if(valorCelda in listaEstadosBusPendienteAhora):
                for i in range(1,sheetExcel.max_column):
                     listaEstadosVerificados.add(row[i].value)
                     listaEstadosBusPendienteNuevos.add(row[i].value)
        listaEstadosBusTerminada=listaEstadosBusTerminada.union(listaEstadosBusPendienteAhora)
        listaEstadosBusPendienteNuevos=listaEstadosBusPendienteNuevos-listaEstadosBusTerminada
        listaEstadosBusPendienteAhora.clear()
        for i in range(0,len(listaEstadosBusPendienteNuevos)):
            listaEstadosBusPendienteAhora.add(listaEstadosBusPendienteNuevos.pop())
        if(len(listaEstadosBusPendienteAhora)==0):
            final=True
    return(listaEstadosVerificados)
              

def automataConvexo(sheetInicial):
    primerEstadoInicial=""
    for row in sheetInicial.iter_rows(min_row=2):
        if(row[0].value!=None):
            if(verprimerEstadoInicial(row[0].value)):
                primerEstadoInicial=row[0].value
                primerEstadoInicial=primerEstadoInicial[1:len(primerEstadoInicial)]
                listaEstadosIniciales.append(primerEstadoInicial)
            elif(verEstadosFinal(row[0].value)):
                estadoFinal=row[0].value
                estadoFinal=estadoFinal[1:len(estadoFinal)]
                listaEstadosFinalesIniciales.append(estadoFinal)
                listaEstadosIniciales.append(estadoFinal)
            else:
                listaEstadosIniciales.append(row[0].value)
                
   # print(primerEstadoInicial)
   # print(listaEstadosFinalesIniciales)
   # print(listaEstadosIniciales)
    return(estadosConvexosNuevo(primerEstadoInicial,listaEstadosFinalesIniciales,listaEstadosIniciales,sheetInicial))

#Esta funcion es para obtener la fila en la que esta un estado
def obtenerFilaEstado(sheetExcel,estado):
    filaEstado=1
    estadoComparador=""
    for row in sheetExcel.iter_rows(min_row=2):
        filaEstado=filaEstado+1
        if(row[0].value!=None):
            estadoComparador=row[0].value
            if(verprimerEstadoInicial(estadoComparador)):
                estadoComparador=estadoComparador[1:len(estadoComparador)]
            elif(verEstadosFinal(estadoComparador)):
                estadoComparador=estadoComparador[1:len(estadoComparador)]
            if(estadoComparador==estado):
                break
    return filaEstado

#Esta funcion se supone es para ver si dos estados son equivalentes pero creo que esta mal
def sonEquivalentesEstados(sheetExcel,estado1,estado2,diccionarioEntrada):
    filaEstado1=obtenerFilaEstado(sheetExcel,estado1)
    filaEstado2=obtenerFilaEstado(sheetExcel,estado2)
    sonEquivalentes=False
    for i in range(2,sheetExcel.max_column+1):
        valorPrueba1=sheet.cell(row=filaEstado1,column=i).value
        valorPrueba2=sheet.cell(row=filaEstado2,column=i).value
        print("dicEva:",diccionarioEntrada," est1:",estado1," est2:",estado2, " valorPr1:",valorPrueba1, "valorPr", valorPrueba2)
        if(diccionarioEntrada[valorPrueba1]==diccionarioEntrada[valorPrueba2]):
            sonEquivalentes=True
        else:
            sonEquivalentes=False
            break
    return sonEquivalentes

#Esta funcion sirve para comparar dos diccionarios que uso como conjunto conciente, pero creo que tampoco funciona
# los diccionario como conjunto conciente los estoy usando de la siguiente forma
# { estado 1: conjunto_al_que pertenece, estado 2: conjunto_al_que pertenece, ..............., estado n:conjunto _al_que_pertenece}          
def compararConjuntosCocientes(conjuntoCociente1,conjuntoConciente2):
    iguales=False
    for estado,conjunto in conjuntoCociente1.items():
        if conjuntoCociente1[estado]==conjuntoConciente2[estado]:
            iguales=True
        else:
            iguales=False
    return iguales        

def conjuntoCocienteCorregido(sheetInicial,listaEstadosConvexos,ListaEstadosFinalesConvexos,estadoInicial):
    diccionarioEstadosConjuntoCociente1=dict(Ejemplo="Inicio")
    diccionarioEstadosConjuntoCociente1.clear()
    diccionarioEstadosConjuntoCociente2=diccionarioEstadosConjuntoCociente1.copy()
    listaEstadosNormales=listaEstadosConvexos-ListaEstadosFinalesConvexos
    for ele in ListaEstadosFinalesConvexos:
        diccionarioEstadosConjuntoCociente1[ele]="0"
    for ele in listaEstadosNormales:
        diccionarioEstadosConjuntoCociente1[ele]="1"
    print(diccionarioEstadosConjuntoCociente1)
    diccionarioEstadosConjuntoCociente2=diccionarioEstadosConjuntoCociente1.copy()
    listaNumeroConjuntos=set(diccionarioEstadosConjuntoCociente1.values())
    contadorNumeroConjuntos=len(listaNumeroConjuntos)
    indiceMaxNumeroConjuntos=contadorNumeroConjuntos-1
    contadorIgualdad=0
    while(contadorIgualdad<=3):
        for conjunto in listaNumeroConjuntos:
            estadosEvaluar=set()
            estadosCambiar=set()
            for estado in listaEstadosConvexos:
                if diccionarioEstadosConjuntoCociente1[estado]==conjunto:
                    estadosEvaluar.add(estado)
            if (len(estadosEvaluar)==2):
                setAux=estadosEvaluar.copy()
                estado1eva=setAux.pop()
                estado2eva=setAux.pop()
                if(sonEquivalentesEstados(sheetInicial,estado1eva,estado2eva,diccionarioEstadosConjuntoCociente1)==False):
                    estadosCambiar.add(estado1eva)
            else:
                for estado1eva in estadosEvaluar:
                    for estado2eva in estadosEvaluar:
                        if estado1eva!=estado2eva:
                            if(sonEquivalentesEstados(sheetInicial,estado1eva,estado2eva,diccionarioEstadosConjuntoCociente1)):
                                estadosCambiar.add(estado1eva)
                                estadosCambiar.add(estado2eva)
           
            if(len(estadosCambiar)==0):
                pass
            else:
                indiceMaxNumeroConjuntos=indiceMaxNumeroConjuntos+1
                for est,con in diccionarioEstadosConjuntoCociente1.items():
                    if (est in estadosCambiar):
                        diccionarioEstadosConjuntoCociente2[est]=str(indiceMaxNumeroConjuntos)
        if (compararConjuntosCocientes(diccionarioEstadosConjuntoCociente1,diccionarioEstadosConjuntoCociente2)):
            contadorIgualdad=contadorIgualdad+1
            diccionarioEstadosConjuntoCociente1=diccionarioEstadosConjuntoCociente2.copy()
        else:
            contadorIgualdad=0
            diccionarioEstadosConjuntoCociente1=diccionarioEstadosConjuntoCociente2.copy()
            listaNumeroConjuntos=set(diccionarioEstadosConjuntoCociente1.values())
            contadorNumeroConjuntos=len(listaNumeroConjuntos)
            indiceMaxNumeroConjuntos=contadorNumeroConjuntos-1
    print("dict",diccionarioEstadosConjuntoCociente1)

#creo que esta funcion tampoco sirve o no se usa
def conjuntoCociente(sheetInicial,listaEstadosConvexos,ListaEstadosFinalesConvexos,estadoInicial):
    diccionarioEstadosConjuntoCociente1=dict(Ejemplo="Inicio")
    diccionarioEstadosConjuntoCociente1.clear()
    diccionarioEstadosConjuntoCociente2=diccionarioEstadosConjuntoCociente1.copy()
    listaEstadosNormales=listaEstadosConvexos-ListaEstadosFinalesConvexos
    for ele in listaEstadosFinalesConvexo:
        diccionarioEstadosConjuntoCociente1[ele]="0"
    for ele in listaEstadosNormales:
        diccionarioEstadosConjuntoCociente1[ele]="1"
    diccionarioEstadosConjuntoCociente2=diccionarioEstadosConjuntoCociente1.copy()
    listaNumeroConjuntos=set(diccionarioEstadosConjuntoCociente1.values())
    contadorNumeroConjuntos=len(listaNumeroConjuntos)
    indiceMaxNumeroConjuntos=contadorNumeroConjuntos-1
    contadorIgualdad=0
    while(contadorIgualdad<=3):
        for conjunto in listaNumeroConjuntos:
            estadosEvaluar=set()
            estadosCambiar=set()
            for estado in listaEstadosConvexos:
                if diccionarioEstadosConjuntoCociente1[estado]==conjunto:
                    estadosEvaluar.add(estado)
            for estado1eva in estadosEvaluar:
                for estado2eva in estadosEvaluar:
                    if estado1eva!=estado2eva:
                        if(sonEquivalentesEstados(sheetInicial,estado1eva,estado2eva,diccionarioEstadosConjuntoCociente1)):
                            print(sonEquivalentesEstados(sheetInicial,estado1eva,estado2eva,diccionarioEstadosConjuntoCociente1))
                            estadosCambiar.add(estado1eva)
                            estadosCambiar.add(estado2eva)
            if(len(estadosCambiar)==0):
                pass
            else:
                indiceMaxNumeroConjuntos=indiceMaxNumeroConjuntos+1
                for est,con in diccionarioEstadosConjuntoCociente1.items():
                    if (est in estadosCambiar):
                        diccionarioEstadosConjuntoCociente2[est]=str(indiceMaxNumeroConjuntos)
        print("diccio1:",diccionarioEstadosConjuntoCociente1," diccio2:",diccionarioEstadosConjuntoCociente2)
        print("COMPARADOdIC",compararConjuntosCocientes(diccionarioEstadosConjuntoCociente1,diccionarioEstadosConjuntoCociente2))
        if (compararConjuntosCocientes(diccionarioEstadosConjuntoCociente1,diccionarioEstadosConjuntoCociente2)):
            contadorIgualdad=contadorIgualdad+1
            diccionarioEstadosConjuntoCociente1=diccionarioEstadosConjuntoCociente2.copy()
        else:
            contadorIgualdad=0
            diccionarioEstadosConjuntoCociente1=diccionarioEstadosConjuntoCociente2.copy()
            listaNumeroConjuntos=set(diccionarioEstadosConjuntoCociente1.values())
            contadorNumeroConjuntos=len(listaNumeroConjuntos)
            indiceMaxNumeroConjuntos=contadorNumeroConjuntos-1
    print("dict",diccionarioEstadosConjuntoCociente1)
        
        


#Menu
FILEPATHUSUARIO=input("Escriba el la direccion del archivo excel:\n")# Esta es la direccion del archivo excel
SHEETUSUARIO=input("Escriba el nombre de la hoja de calculo en excel:\n")#Esta es la hoja de calculo de excel
workbook=load_workbook(FILEPATHUSUARIO)#Se crea el objeto para trabajar con excel
sheet=workbook[SHEETUSUARIO]#se crea el objeto para trabajar con la hoja de calculo de excel
listaEstadosAutomataConvexo=automataConvexo(sheet)# Se dan los estados del automata convexo
setListaEstadosIniciales=set(listaEstadosIniciales)# Se crean los estados del automata antes de ser convexo
listaEstadosDesechadosConvexo=setListaEstadosIniciales-listaEstadosAutomataConvexo# se dan los estados que no quedaron en el automata convexo
listaEstadosFinalesConvexo=set(listaEstadosFinalesIniciales)-listaEstadosDesechadosConvexo#se dan los estados finales del auotomata convexo
#print(listaEstadosAutomataConvexo)
conjuntoCocienteCorregido(sheet,listaEstadosAutomataConvexo,listaEstadosFinalesConvexo,primerEstadoInicial)#Aui se trata de crear el conjunto cociente
